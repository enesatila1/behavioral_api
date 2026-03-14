from typing import List, Optional, TYPE_CHECKING
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from datalayer.repository import TrialResultRepository, ParticipantRepository
from datalayer.model.db import TrialResult

if TYPE_CHECKING:
    from firebase_admin.firestore_async import AsyncClient


class ExportService:
    """Service for exporting test results to Excel format"""

    def __init__(self, db: "AsyncClient"):
        self.db = db
        self.participant_repo = ParticipantRepository(db)
        self.trial_repo = TrialResultRepository(db)

    async def export_to_excel(self, participant_id: Optional[str] = None) -> bytes:
        """
        Export test results to Excel format.

        Args:
            participant_id: If provided, export only this participant's results.
                           If None, export all results.

        Returns:
            bytes: Excel file content
        """
        try:
            if participant_id:
                trials = await self.trial_repo.find_by_participant_id(participant_id)
            else:
                trials = await self.trial_repo.get_all()

            return self._generate_excel(trials)
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            raise

    def _generate_excel(self, trials: List[TrialResult]) -> bytes:
        """Generate Excel workbook from trial results"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Define columns
        columns = [
            "Participant ID",
            "Participant Name",
            "Test Type",
            "Trial Index",
            "RT (ms)",
            "Response",
            "Task Type",
            "Sentence ID",
            "Word",
            "Word Position",
            "Condition",
            "Sentence",
            "Is Grammatical",
            "Attrition Marker",
            "Accuracy",
            "Correct Answer",
            "Created At",
        ]

        # Add header row with styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_num, trial in enumerate(trials, 2):
            row_data = [
                trial.participant_id,
                trial.participant_name,
                trial.test_type,
                trial.trial_index,
                trial.rt,
                trial.response,
                trial.task_type,
                trial.sentence_id,
                trial.word,
                trial.word_position,
                trial.condition,
                trial.sentence,
                trial.is_grammatical,
                trial.attrition_marker,
                trial.accuracy,
                trial.correct_answer,
                trial.created_at.isoformat() if trial.created_at else None,
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
